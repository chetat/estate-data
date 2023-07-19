import unittest
from datetime import datetime
from unittest.mock import MagicMock, Mock, patch

from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app import create_app
from app import sqlalchemy as db
from app.config import TestingConfig
from app.models import Tenant, TenantTransaction
from app.repository.transaction import TenantTransactionRepository
from main import (
    convert_tenant_row_to_dict,
    load_data_to_db,
    load_workbook_data,
    prepare_tenant_item,
    process_transaction,
    save_tenant_data,
)


class TestLoadWorkbookData(unittest.TestCase):
    """
    Test class for load_workbook_data() function.
    """

    @patch("main.load_workbook")
    def test_load_workbook_data_valid_file(self, mock_load_workbook):
        """
        Test case for valid excel file.
        """
        mock_workbook = MagicMock()
        mock_worksheet = MagicMock()
        mock_workbook.worksheets = [mock_worksheet]
        mock_load_workbook.return_value = mock_workbook
        mock_worksheet.rows = [[MagicMock(value=i) for i in range(3)] for _ in range(3)]

        result = load_workbook_data("valid_file.xlsx")
        expected_result = [[i for i in range(3)] for _ in range(3)]

        self.assertEqual(result, expected_result)

    @patch("main.load_workbook")
    def test_load_workbook_data_invalid_file(self, mock_load_workbook):
        """
        Test case for invalid excel file.
        """
        mock_load_workbook.side_effect = InvalidFileException
        result = load_workbook_data("invalid_file.xlsx")
        self.assertIsNone(result)

    @patch("main.load_workbook")
    def test_load_workbook_data_unexpected_error(self, mock_load_workbook):
        """
        Test case for unexpected error while loading excel file.
        """
        mock_load_workbook.side_effect = Exception
        result = load_workbook_data("unexpected_error.xlsx")
        self.assertIsNone(result)


class TestTenantDataConversion(unittest.TestCase):
    """
    Test class for functions converting tenant data.
    """

    def test_convert_tenant_row_to_dict(self):
        """
        Test case for conversion of tenant row to dictionary.
        """
        row = [
            "Tenant",
            "yeku",
            "Code",
            "1234",
            "Vacate",
            None,
            "Main Unit No",
            "1234",
            "Property",
            "1234",
            "General Contact",
            "1234",
            "Telephone",
            "1234",
            "Lease Starts",
            "01/01/2022",
            "Ends",
            "04/01/2022",
        ]
        result = convert_tenant_row_to_dict(row)
        expected = {
            "Tenant": "yeku",
            "Code": "1234",
            "Vacate": None,
            "Main Unit No": "1234",
            "Property": "1234",
            "General Contact": "1234",
            "Telephone": "1234",
            "Lease Starts": "01/01/2022",
            "Ends": "04/01/2022",
        }
        self.assertEqual(result, expected)

    def test_prepare_tenant_item(self):
        """
        Test case for preparing tenant item.
        """
        row = [
            "Tenant",
            "yeku",
            "Code",
            "1234",
            "Vacate",
            None,
            "Main Unit No",
            "1234",
            "Property",
            "1234",
            "General Contact",
            "1234",
            "Telephone",
            "1234",
            "Lease Starts",
            "01/01/2022",
            "Ends",
            "04/01/2022",
        ]
        result = prepare_tenant_item(row)
        expected = Tenant(
            tenant_name="yeku",
            tenant_code="1234",
            tenant_is_active=True,
            tenant_main_unit_no="1234",
            tenant_property_name="1234",
            tenant_general_contact="1234",
            tenant_telephone="1234",
            tenant_lease_start_date="1234",
            tenant_lease_end_date="1234",
            tenant_vacate_date=None,
        )
        self.assertEqual(result.tenant_name, expected.tenant_name)


class TestTransactionProcessing(unittest.TestCase):
    """
    Test class for process_transaction() function.
    """

    @patch("app.repository.transaction.TenantTransactionRepository")
    def test_process_transaction(self, mock_transaction_repo):
        """
        Test case for processing a transaction.
        """
        tenant_transaction_info = {
            "Period": "1",
            "Date": "2022-01-01",
            "Transaction": "transaction",
            "Description": "description",
            "Tax": "10",
            "Remarks": "remarks",
            "Exclusive": "100",
            "Inclusive": "110",
        }
        tenant_id = 1
        last_inserted_period = "1"
        process_transaction(
            tenant_transaction_info,
            tenant_id,
            last_inserted_period,
            mock_transaction_repo,
        )
        mock_transaction_repo.add.assert_called()


class TestSaveTenantData(unittest.TestCase):
    """
    Test class for save_tenant_data() function.
    """

    def setUp(self):
        """
        Set up context and database for testing.
        """
        self.app = create_app(TestingConfig())
        self.app_context = self.app.app_context()
        self.app_context.push()

        # binds the app to the current context
        with self.app.app_context():
            self.db = db
            self.db.create_all()

    def tearDown(self):
        """
        Clean up context and database after testing.
        """
        self.db.session.remove()
        self.db.drop_all()
        self.app_context.pop()

    @patch.object(TenantTransactionRepository, "bulk_add")
    def test_save_tenant_data(self, mock_bulk_add):
        """
        Test case for successful save of tenant data.
        """
        db_session = Session()
        db_session.commit = Mock()
        db_session.rollback = Mock()
        mock_transaction_repo = TenantTransactionRepository(db_session)

        prepared_tenant_data = {1: {"transactions": ["transaction1"]}}
        save_tenant_data(mock_transaction_repo, prepared_tenant_data, db_session)

        self.assertEqual(
            mock_bulk_add.call_count, len(prepared_tenant_data[1]["transactions"])
        )
        db_session.commit.assert_called_once()

    @patch.object(TenantTransactionRepository, "bulk_add")
    def test_save_tenant_data_with_exception(self, mock_bulk_add):
        """
        Test case for unsuccessful save of tenant data.
        """
        db_session = Session()
        db_session.commit = Mock()
        db_session.rollback = Mock()
        mock_bulk_add.side_effect = Exception("Something went wrong")
        mock_transaction_repo = TenantTransactionRepository(db_session)
        prepared_tenant_data = {1: {"transactions": [TenantTransaction()]}}
        with self.assertRaises(Exception):
            save_tenant_data(mock_transaction_repo, prepared_tenant_data)
        # Verify rollback worked by checking no entries were added to the database
        self.assertEqual(db.session.query(TenantTransaction).count(), 0)


class TestLoadDataToDB(unittest.TestCase):
    """
    Test class for load_data_to_db() function.
    """

    def setUp(self):
        """
        Set up context and database for testing.
        """
        self.app = create_app(TestingConfig())
        self.app_context = self.app.app_context()
        self.app_context.push()

        # binds the app to the current context
        with self.app.app_context():
            self.db = db
            # create all tables
            self.db.create_all()

    def tearDown(self):
        """
        Clean up context and database after testing.
        """
        self.db.session.remove()
        self.db.drop_all()
        self.app_context.pop()

    @patch("main.load_workbook_data")
    @patch("main.TenantRepository")
    @patch("main.TenantTransactionRepository")
    def test_load_data_to_db(
        self, mock_transaction_repo, mock_tenant_repo, mock_load_workbook_data
    ):
        """
        Test case for loading data to database.
        """
        lease_start = datetime.utcnow().date()
        lease_end = datetime.utcnow().date()
        mock_load_workbook_data.return_value = [
            [
                "Tenant",
                "yeku",
                "Code",
                "1234",
                "Vacate",
                None,
                "Main Unit No",
                "1234",
                "Property",
                "1234",
                "General Contact",
                "1234",
                "Telephone",
                "1234",
                "Lease Starts",
                lease_start,
                "Ends",
                lease_end,
            ],
            [
                "1",
                "2022-01-01",
                "transaction",
                "description",
                "10",
                "remarks",
                "100",
                "110",
                "200",
            ],
        ]
        mock_tenant = mock_tenant_repo.return_value.add.return_value
        mock_tenant.id = 1
        mock_transction = mock_transaction_repo.return_value.add.return_value
        mock_transction.id = 1
        load_data_to_db()

        mock_tenant_repo.assert_called_once()
        mock_transaction_repo.assert_called_once()

        targs, _ = mock_tenant_repo.return_value.add.call_args
        created_tenant = targs[0]
        self.assertEqual(created_tenant.tenant_name, "yeku")
        self.assertEqual(created_tenant.tenant_code, "1234")

        # Checking if the method was called
        if mock_transaction_repo.return_value.add.called: 
            tx_args, _ = mock_transaction_repo.return_value.add.call_args
            created_transaction = tx_args[0]
            print(tx_args)
            self.assertEqual(created_transaction.period, 1)
            self.assertEqual(created_transaction.date, datetime(2022, 1, 1).date())
            self.assertEqual(created_transaction.transaction, "transaction")

if __name__ == "__main__":
    unittest.main()
