from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app import create_app
from app import sqlalchemy as db
from app.models import Tenant, TenantTransaction
from app.repository.tenants import TenantRepository
from app.repository.transaction import TenantTransactionRepository


def load_workbook_data(filename):
    try:
        workbook = load_workbook(filename=filename)
        worksheet = workbook.worksheets[0]
        row_list = [[cell.value for cell in row] for row in worksheet.rows]
        return row_list
    except InvalidFileException:
        print(
            f"Invalid file: {filename}.Please make sure the file exists"
        )
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


def convert_tenant_row_to_dict(row):
    result = {}
    for i in range(0, len(row), 2):
        key = row[i].strip(",")
        value = row[i + 1] if i + 1 < len(row) else None
        result[key] = value
    return result


def prepare_tenant_item(row: list):
    tenant_info = convert_tenant_row_to_dict(row)
    vacate_date = tenant_info.get("Vacate")
    tenant_is_active = True if not vacate_date else False
    tenant_info["tenant_is_active"] = tenant_is_active
    tenant_model = {
        "tenant_name": tenant_info["Tenant"],
        "tenant_code": tenant_info["Code"],
        "tenant_main_unit_no": tenant_info["Main Unit No"],
        "tenant_property_name": tenant_info["Property"],
        "tenant_general_contact": tenant_info.get("General Contact", None),
        "tenant_telephone": tenant_info["Telephone"],
        "tenant_lease_start_date": tenant_info.get("Lease Starts", None),
        "tenant_lease_end_date": tenant_info["Ends"],
        "tenant_vacate_date": vacate_date,
        "tenant_is_active": tenant_is_active,
    }
    tenant = Tenant(**tenant_model)
    return tenant


def process_transaction(
    tenant_transaction_info: dict,
    tenant_id: int,
    last_inserted_period: str,
    transaction_repo: TenantTransactionRepository,
):
    period = tenant_transaction_info["Period"]
    date = tenant_transaction_info["Date"]
    transaction = tenant_transaction_info["Transaction"]
    description = tenant_transaction_info["Description"]
    tax = tenant_transaction_info["Tax"]
    remarks = tenant_transaction_info["Remarks"]
    exclusive = tenant_transaction_info["Exclusive"]
    inclusive = tenant_transaction_info["Inclusive"]
    if not period:
        period = last_inserted_period
    last_inserted_period = period

    if not inclusive or inclusive == "0.00":
        return None

    tenant_transaction = TenantTransaction(
        period=int(period),
        date=date,
        transaction=transaction,
        tax=tax,
        remarks=remarks,
        exclusive=exclusive,
        inclusive=inclusive,
        description=description,
        tenant_id=tenant_id,
    )
    transaction_repo.add(tenant_transaction)


def save_tenant_data(
    tenant_transaction_repo: TenantTransactionRepository,
    prepared_tenant_data: dict,
    db: Session,
):
    for tenant_id, tenant_data in prepared_tenant_data.items():
        try:
            tenant_transaction_repo.bulk_add(tenant_data["transactions"])
            db.commit()
        except Exception as exc:
            db.rollback()
            print(f"Error occurred for Tenant_id {tenant_id}: {exc}")
            raise exc


# write function to read csv file from directory and loop through the rows
def load_data_to_db():
    tenant_repo = TenantRepository(db.session)
    transaction_repo = TenantTransactionRepository(db.session)
    row_list = load_workbook_data(filename="data/data.xlsx")
    checked_tenant_ids = []
    current_tenant_id = None
    # I assume the first cell of period column will always be filled with a value
    last_inserted_period = None
    prepared_data = {}
    for row in row_list:
        if row[0] == "Tenant":
            tenant = prepare_tenant_item(row)
            tenant_repo.add(tenant)
            current_tenant_id = tenant.id
            checked_tenant_ids = []
            checked_tenant_ids.append(current_tenant_id)
            prepared_data[current_tenant_id] = {"transactions": []}
        elif current_tenant_id in checked_tenant_ids:
            # What happens if period is not filled in the first transaction row?
            period = row[0]
            date = row[1]
            transaction = row[2]
            description = row[3]
            tax = float(row[4]) if row[4] else 0
            remarks = row[5]
            exclusive = float(row[6]) if row[6] else 0
            inclusive = row[8]
            if not period:
                period = last_inserted_period
            last_inserted_period = period

            if not inclusive or inclusive == "0.00":
                continue

            tenant_transaction = TenantTransaction(
                period=int(period),
                date=date,
                transaction=transaction,
                tax=tax,
                remarks=remarks,
                exclusive=exclusive,
                inclusive=inclusive,
                description=description,
                tenant_id=current_tenant_id
            )
            prepared_data[current_tenant_id]["transactions"].append(tenant_transaction)
        else:
            continue
    save_tenant_data(transaction_repo, prepared_data, db.session)


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        load_data_to_db()
